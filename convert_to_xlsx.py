import xml.etree.ElementTree as ET
import re

filepath = r'C:\QNBFuelExtension\QNB_Fuel_3Sheets_01072023_to_30062026 (2).xls'
output = r'C:\QNBFuelExtension\QNB_Fuel_3Sheets_01072023_to_30062026_fixed.xlsx'

with open(filepath, 'r', encoding='utf-8') as f:
    content = f.read()

# Fix <Cell ss:StyleID="X"}> -> <Cell ss:StyleID="X">
content = re.sub(r'<Cell ss:StyleID="[^"]*"}>', '<Cell>', content)

# Also fix any other }> patterns
content = re.sub(r'\}>', '>', content)

print("Trying to parse XML...")
tree = ET.fromstring(content)

namespaces = {
    'ss': 'urn:schemas-microsoft-com:office:spreadsheet'
}

from openpyxl import Workbook

wb = Workbook()
ws = wb.active

for worksheet in tree.findall('.//ss:Worksheet', namespaces):
    sheet_name = worksheet.get('{urn:schemas-microsoft-com:office:spreadsheet}Name')
    print(f"Processing sheet: {sheet_name}")
    
    if sheet_name:
        ws = wb.create_sheet(title=sheet_name[:31])
    else:
        ws = wb.active
    
    row_idx = 1
    
    for row in worksheet.findall('.//ss:Row', namespaces):
        cell_idx = 1
        for cell in row.findall('ss:Cell', namespaces):
            data_elem = cell.find('ss:Data', namespaces)
            if data_elem is not None:
                value = data_elem.text or ''
                ws.cell(row=row_idx, column=cell_idx, value=value)
            cell_idx += 1
        row_idx += 1

wb.save(output)
print(f"Saved as {output}")